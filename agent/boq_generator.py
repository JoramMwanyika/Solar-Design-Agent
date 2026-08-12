"""
BOQ Generator for Solar PV Systems.
Generates structured Bill of Quantities aligned with client sample workbooks:
- 14-15 standard installation & equipment categories
- Strict quantities-only mode ("Pricing in BOQ — should include quantities only")
- Includes exact items: PV modules, mounting structure, solar string cables, inverters/BESS,
  DCDB/ACDB switchgear, earthing/LPS, cable containment/trunking, civil works, installation & commissioning.
- Exports both professional Excel (.xlsx) using openpyxl and clean Markdown tables.
"""
from io import BytesIO
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_boq(sizing: dict, project_name: str = "Solar PV System") -> List[Dict]:
    """
    Generates a standardized BOQ item list from sizing results.
    Strictly outputs quantities only (unit costs / total costs left empty/0 per client spec).
    """
    boq: List[Dict] = []
    
    # Extract key sizing parameters
    panel_qty = sizing.get("panel_qty", 80)
    panel_wp = sizing.get("panel_wp", 625)
    inverter_kw = sizing.get("inverter", {}).get("kw", 50)
    inverter_qty = sizing.get("inverter", {}).get("qty", 1)
    inverter_brand = sizing.get("inverter", {}).get("brand", "")
    voltage_arch = sizing.get("inverter", {}).get("voltage_architecture", "")
    
    battery_info = sizing.get("battery", {})
    battery_qty = battery_info.get("qty", 0)
    battery_type = battery_info.get("type", "Dyness Stack280 14.33kWh HV Battery")
    
    string_info = sizing.get("stringing", {})
    total_strings = string_info.get("total_strings", 6)
    
    cables_info = sizing.get("cables", {})
    dc_sqmm = cables_info.get("dc_sqmm", 4)
    dc_total_m = cables_info.get("dc_total_m", 120)
    ac_sqmm = cables_info.get("ac_sqmm", 25)
    ac_breaker_a = cables_info.get("ac_breaker_a", 125)
    
    system_type = sizing.get("system_type", "hybrid")

    # 1. Solar Panels
    boq.append({
        "item_no": "1",
        "category": "Solar Panels",
        "description": f"{panel_wp}Wp High-Efficiency Monocrystalline PV Modules (Tier 1 Jinko/JA Solar/Longi)",
        "unit": "Pcs",
        "quantity": panel_qty,
        "unit_cost": "",
        "total_cost": "",
        "remarks": f"Total PV Capacity: {(panel_qty * panel_wp)/1000:.2f} kWp"
    })

    # 2. Solar Mounting Structure & Accessories
    boq.append({
        "item_no": "2.1",
        "category": "Mounting Structure",
        "description": "Anodized Aluminum Roof/Carport Mounting Structure Rails, L-Feet, Roof Hooks & Clamps",
        "unit": "Set",
        "quantity": panel_qty,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Includes mid/end clamps, earthing clips, and EPDM waterproof rubber gaskets"
    })
    boq.append({
        "item_no": "2.2",
        "category": "Mounting Structure",
        "description": "Stainless steel fasteners, hanger bolts, and structural rail joiners/connectors",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Corrosion resistant grade A2/A4"
    })

    # 3. Solar String Cables & Connectors
    boq.append({
        "item_no": "3.1",
        "category": "DC Cabling",
        "description": f"Solar DC Cable {dc_sqmm}mm² Red (Tinned Copper, XLPO, 1.5/1.5kVdc UV Resistant)",
        "unit": "m",
        "quantity": max(100, int(dc_total_m * 0.55)),
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Positive string runs from array to PVDB"
    })
    boq.append({
        "item_no": "3.2",
        "category": "DC Cabling",
        "description": f"Solar DC Cable {dc_sqmm}mm² Black (Tinned Copper, XLPO, 1.5/1.5kVdc UV Resistant)",
        "unit": "m",
        "quantity": max(100, int(dc_total_m * 0.55)),
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Negative string runs from array to PVDB"
    })
    boq.append({
        "item_no": "3.3",
        "category": "DC Cabling",
        "description": "MC4 Male/Female Waterproof Branch & String Connectors (1500V DC rated, IP68)",
        "unit": "Pair",
        "quantity": max(10, total_strings * 3),
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Original Staubli or certified equivalent"
    })

    # 4. Inverters & BESS
    if inverter_brand:
        inv_desc = f"{inverter_brand} ({inverter_kw:.0f}kW 3-Phase Solar Inverter)"
    else:
        inv_desc = f"{inverter_kw}kW On-Grid / Hybrid 3-Phase Solar Inverter (Deye / Huawei / GoodWe / Sungrow)"
    boq.append({
        "item_no": "4.1",
        "category": "Power Conversion",
        "description": inv_desc,
        "unit": "Pcs",
        "quantity": inverter_qty,
        "unit_cost": "",
        "total_cost": "",
        "remarks": f"{voltage_arch} — Includes Wi-Fi/4G smart dongle and CT current sensors" if voltage_arch else "Includes Wi-Fi/4G smart dongle and CT current sensors for zero-export"
    })

    if system_type in ("off-grid", "hybrid") and battery_qty > 0:
        boq.append({
            "item_no": "4.2",
            "category": "Battery Storage (BESS)",
            "description": f"{battery_type} (High Voltage LFP Battery Stack)",
            "unit": "Pcs",
            "quantity": battery_qty,
            "unit_cost": "",
            "total_cost": "",
            "remarks": f"Total BESS capacity: {battery_qty * 14.33:.1f} kWh"
        })
        boq.append({
            "item_no": "4.3",
            "category": "Battery Storage (BESS)",
            "description": "High Voltage Battery Rack / Stack Enclosure, BMS Controller & Communication cables",
            "unit": "Set",
            "quantity": max(1, (battery_qty + 8) // 9),
            "unit_cost": "",
            "total_cost": "",
            "remarks": "Includes HV BMS control box and base pedestal"
        })

    # 5. DC Protection & Switchgear
    boq.append({
        "item_no": "5.1",
        "category": "DC Switchgear",
        "description": f"PV Combiner Box / DC Distribution Board (PVDB) for {total_strings} Strings with 1000V DC SPD Type II, Fuses & Isolator",
        "unit": "Set",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "IP65 weatherproof polycarbonate wall-mounted enclosure"
    })
    if battery_qty > 0:
        boq.append({
            "item_no": "5.2",
            "category": "DC Switchgear",
            "description": "Battery Breaker Box with 160A/250A 1000V DC Molded Case Circuit Breaker (MCCB)",
            "unit": "Set",
            "quantity": 1,
            "unit_cost": "",
            "total_cost": "",
            "remarks": "Sized with 1.25x safety margin for high charge/discharge current"
        })

    # 6. AC Switchgear & Board
    boq.append({
        "item_no": "6",
        "category": "AC Switchgear",
        "description": f"AC Distribution Board (ACDB) with {ac_breaker_a}A 4-Pole MCCB, Type II AC SPD, Over/Under Voltage Relay & Pilot Lights",
        "unit": "Set",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Connects inverter output to main facility supply/grid panel"
    })

    # 7. DC & AC Cables / Feeder Runs
    if battery_qty > 0:
        boq.append({
            "item_no": "7.1",
            "category": "Cabling & Feeder Runs",
            "description": "Battery HV Power Cable 50mm² / 16mm² CU/PVC/PVC-Nitrile (Tower to Breaker & Inverter)",
            "unit": "m",
            "quantity": 30,
            "unit_cost": "",
            "total_cost": "",
            "remarks": "Flexible orange/black battery power runs with heavy duty lugs"
        })
    boq.append({
        "item_no": "7.2",
        "category": "Cabling & Feeder Runs",
        "description": f"4-Core {ac_sqmm}mm² CU/XLPE/PVC Armoured/Unarmoured AC Main Feeder Cable (0.6/1kV)",
        "unit": "m",
        "quantity": max(50, int(sizing.get("cables", {}).get("ac_distance_m", 60))),
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Main AC interconnection between inverter and facility switchboard"
    })
    boq.append({
        "item_no": "7.3",
        "category": "Cabling & Feeder Runs",
        "description": "Heavy duty tinned copper cable lugs, cable ties, glands, heat shrink tubes, and warning labels",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Complete electrical termination kit"
    })

    # 8. Earthing & Lightning Protection System (LPS)
    boq.append({
        "item_no": "8.1",
        "category": "Earthing & LPS",
        "description": "Earthing Cable 16mm² CU/PVC Green/Yellow (Inverter, PVDB, Roof to Main Earthpit)",
        "unit": "m",
        "quantity": 120,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Dedicated clean earth and structure bonding runs"
    })
    boq.append({
        "item_no": "8.2",
        "category": "Earthing & LPS",
        "description": "Earthing Cable 6mm² CU/PVC Green/Yellow (Rail-to-Rail inter-bonding)",
        "unit": "m",
        "quantity": 60,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Equipotential bonding of all solar array rails"
    })
    boq.append({
        "item_no": "8.3",
        "category": "Earthing & LPS",
        "description": "Pure Copper Earth Rod (1.5m x 16mm) with heavy duty brass clamp & inspection pit cover",
        "unit": "Set",
        "quantity": 3,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Achieving earth resistance < 5 ohms"
    })
    boq.append({
        "item_no": "8.4",
        "category": "Earthing & LPS",
        "description": "25x3mm Copper Tape / Air terminal system for Lightning Protection System (LPS)",
        "unit": "m",
        "quantity": 40,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Surge and direct strike mitigation"
    })

    # 9. Containment & Cable Trunking
    boq.append({
        "item_no": "9.1",
        "category": "Containment",
        "description": "Hot-dip galvanized perforated cable tray (150x50mm) with covers and jointing brackets",
        "unit": "m",
        "quantity": 45,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Outdoor / rooftop cable protection"
    })
    boq.append({
        "item_no": "9.2",
        "category": "Containment",
        "description": "Heavy duty PVC trunking and flexible conduit pipes (25mm/32mm) with saddles and fittings",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Indoor wall-mounted and drop cable runs"
    })

    # 10. Civil Works & Foundations
    boq.append({
        "item_no": "10.1",
        "category": "Civil Works",
        "description": "Reinforced concrete plinth / equipment foundation for inverter, BESS rack and outdoor enclosures",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Grade C25/30 concrete with cable trenching if outdoor"
    })
    boq.append({
        "item_no": "10.2",
        "category": "Civil Works",
        "description": "Earth pit excavation, chemical treatment (Bentonite/Marconite) and backfilling",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "For grounding system installation"
    })

    # 11. Installation & Electrical Assembly
    boq.append({
        "item_no": "11.1",
        "category": "Installation Services",
        "description": "Mechanical installation of mounting structures and alignment/clamping of solar modules",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Certified mechanical solar installation crew"
    })
    boq.append({
        "item_no": "11.2",
        "category": "Installation Services",
        "description": "Complete electrical wiring, cable pulling, termination, switchgear mounting and interconnections",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Licensed electrical technicians per local codes"
    })

    # 12. Testing, Commissioning & Documentation
    boq.append({
        "item_no": "12.1",
        "category": "Testing & Commissioning",
        "description": "System testing (Voc, Isc, insulation resistance, earth resistance), inverter programming and grid commissioning",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Full handover checklist with performance verification"
    })
    boq.append({
        "item_no": "12.2",
        "category": "Testing & Commissioning",
        "description": "As-Built electrical single-line drawings (SLD), operation & maintenance manuals, and user training",
        "unit": "Set",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Includes factory warranty certificates for panels/inverters/batteries"
    })

    # 13. Smart Monitoring System
    boq.append({
        "item_no": "13",
        "category": "Monitoring",
        "description": "Smart Energy Meter with 3-Phase CT coils for real-time generation, load tracking & remote portal access",
        "unit": "Set",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Includes lifetime cloud monitoring account setup"
    })

    # 14. Transportation & Site Logistics
    boq.append({
        "item_no": "14",
        "category": "Logistics",
        "description": "Transportation, insurance in transit, offloading and site material handling for all equipment",
        "unit": "Lot",
        "quantity": 1,
        "unit_cost": "",
        "total_cost": "",
        "remarks": "Delivery directly to project site"
    })

    return boq


def boq_to_markdown_table(boq_items: List[Dict]) -> str:
    """Converts BOQ items list into a clean markdown table (Quantities only mode)."""
    lines = [
        "| Item | Category | Description & Specifications | Qty | Unit | Unit Cost | Total Cost | Remarks |",
        "|---|---|---|:---:|:---:|:---:|:---:|---|",
    ]
    for item in boq_items:
        lines.append(
            f"| `{item.get('item_no', '')}` | **{item.get('category', '')}** | {item.get('description', '')} | "
            f"**{item.get('quantity', '')}** | `{item.get('unit', '')}` | - | - | *{item.get('remarks', '')}* |"
        )
    return "\n".join(lines)


def boq_to_markdown(boq: List[Dict], project_name: str = "Solar PV System") -> str:
    """Convenience wrapper around boq_to_markdown_table with header."""
    return f"## 📋 Bill of Quantities (BOQ) — {project_name}\n\n" + boq_to_markdown_table(boq)


def generate_boq_excel(
    boq_items: List[Dict],
    project_name: str = "Solar PV System",
    system_type: str = "hybrid",
    location: str = "",
    client_name: str = "",
    prepared_by: str = "",
    sizing_summary: Optional[Dict] = None,
) -> bytes:
    """
    Generates a professional, beautifully styled Excel spreadsheet (.xlsx) for the BOQ.
    Strict quantities-only pricing rule is applied (Unit Cost and Total Cost columns are left open for procurement).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"

    # Fonts & Fills
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Navy Blue
    category_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid") # Soft Blue/Grey
    price_col_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid") # Soft Salmon per client template

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = f"BILL OF QUANTITIES — {project_name.upper()}"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Metadata Header
    ws["A3"] = "Project Name:"
    ws["B3"] = project_name
    ws["A4"] = "Client Name:"
    ws["B4"] = client_name or "General Client"
    ws["E3"] = "System Type:"
    ws["F3"] = system_type.upper()
    ws["E4"] = "Location:"
    ws["F4"] = location or "East Africa"
    ws["E5"] = "Prepared By:"
    ws["F5"] = prepared_by or "Solar Design Agent AI"
    
    for row in range(3, 6):
        for col in [1, 5]:
            ws.cell(row=row, column=col).font = bold_font

    # Table Headers
    headers = ["Item No.", "Category", "Description & Specifications", "Qty", "Unit", "Rate / Unit Cost", "Total Cost", "Remarks"]
    header_row = 7
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    # Insert Items
    curr_row = 8
    curr_category = ""
    for item in boq_items:
        # Category separator row if new category
        cat = item.get("category", "")
        if cat != curr_category:
            curr_category = cat
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            cat_cell = ws.cell(row=curr_row, column=1, value=f"  {curr_category.upper()}")
            cat_cell.font = bold_font
            cat_cell.fill = category_fill
            cat_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[curr_row].height = 22
            for c in range(1, 9):
                ws.cell(row=curr_row, column=c).border = thin_border
            curr_row += 1

        # Item row
        ws.cell(row=curr_row, column=1, value=item.get("item_no", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=2, value=cat).alignment = Alignment(vertical="center")
        desc_cell = ws.cell(row=curr_row, column=3, value=item.get("description", ""))
        desc_cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        ws.cell(row=curr_row, column=4, value=item.get("quantity", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=5, value=item.get("unit", "")).alignment = Alignment(horizontal="center", vertical="center")
        
        # Unit Cost & Total Cost columns left blank with soft salmon fill ("Pricing in BOQ — should include quantities only")
        uc_cell = ws.cell(row=curr_row, column=6, value="")
        uc_cell.fill = price_col_fill
        tc_cell = ws.cell(row=curr_row, column=7, value="")
        tc_cell.fill = price_col_fill
        
        ws.cell(row=curr_row, column=8, value=item.get("remarks", "")).alignment = Alignment(vertical="center")

        for c in range(1, 9):
            ws.cell(row=curr_row, column=c).font = data_font
            ws.cell(row=curr_row, column=c).border = thin_border
            
        ws.row_dimensions[curr_row].height = 24
        curr_row += 1

    # Summary Row
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
    sum_cell = ws.cell(row=curr_row, column=1, value="SUB-TOTAL FOR EQUIPMENT & INSTALLATION (QUANTITIES ONLY)")
    sum_cell.font = bold_font
    sum_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=6).fill = price_col_fill
    ws.cell(row=curr_row, column=7).fill = price_col_fill
    ws.row_dimensions[curr_row].height = 26
    for c in range(1, 9):
        ws.cell(row=curr_row, column=c).border = thin_border

    # Auto-adjust column widths
    col_widths = {1: 10, 2: 22, 3: 45, 4: 10, 5: 10, 6: 16, 7: 16, 8: 35}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_sizing_and_design_workbook(
    sizing: dict,
    boq_items: Optional[List[Dict]] = None,
    project_name: str = "Solar PV System Design Workbook",
    location: str = "East Africa",
    client_name: str = "",
    prepared_by: str = "",
) -> bytes:
    """
    Generates a comprehensive, multi-sheet engineering sizing and design workbook (.xlsx).
    Includes:
      Sheet 1: Executive Summary & Load Demand
      Sheet 2: PV Array & Stringing Configuration
      Sheet 3: Battery Storage System (BESS) & Protection
      Sheet 4: Inverter Selection & AC Switchgear
      Sheet 5: Cable Sizing (DC String / AC Feeder) & LPS Schedule
      Sheet 6: Complete Bill of Quantities (Quantities Only)
    """
    if not boq_items:
        boq_items = generate_boq(sizing, project_name)

    wb = openpyxl.Workbook()
    
    # Styles
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    data_font = Font(name="Calibri", size=11)
    
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    def format_title_banner(ws, title_text: str, cols: int = 5):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        cell = ws.cell(row=1, column=1, value=f"  {title_text.upper()} — {project_name.upper()}")
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 36
        
        ws.cell(row=2, column=1, value=f"Project: {project_name} | Location: {location} | Prepared By: {prepared_by or 'SolarBot AI'} | System Type: {sizing.get('system_type', 'Hybrid').upper()}")
        ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="595959")
        ws.row_dimensions[2].height = 20

    def write_table(ws, start_row: int, headers: List[str], rows: List[List[str]], col_widths: Dict[int, int]):
        # Write headers
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        ws.row_dimensions[start_row].height = 24
        
        # Write data rows
        curr = start_row + 1
        for i, row_data in enumerate(rows):
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=curr, column=col_idx, value=val)
                c.font = data_font
                c.border = thin_border
                if i % 2 == 1:
                    c.fill = zebra_fill
                if col_idx == 1:
                    c.font = bold_font
                elif col_idx in (2, 3):
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[curr].height = 22
            curr += 1
        
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        return curr + 1

    # ── Sheet 1: Executive Summary & Load Demand ──
    ws1 = wb.active
    ws1.title = "1. Executive Summary"
    format_title_banner(ws1, "System Executive Summary & Load Demand", cols=4)
    
    sys_type = sizing.get("system_type", "hybrid")
    sizing_basis = "Apparent Power (S) in kVA" if sys_type in ("off-grid", "hybrid") else "Active Power (P) in kW"
    
    summary_rows = [
        ["System Classification", sys_type.upper(), "Standard off-grid / hybrid / grid-tied configuration"],
        ["Location & Solar Resource", f"{location} ({sizing.get('peak_sun_hours', 5.0)} Peak Sun Hours)", "Based on regional solar radiation data"],
        ["Inverter Sizing Basis", sizing_basis, "IEC 62548 & AS/NZS 4509 engineering rule"],
        ["Peak Active Power Demand (P)", f"{sizing.get('total_peak_power_w', 0):,.1f} W ({sizing.get('total_peak_power_w', 0)/1000:,.2f} kW)", "Total connected load real power draw"],
        ["Peak Apparent Power Demand (S)", f"{sizing.get('total_peak_va', 0):,.1f} VA ({sizing.get('total_peak_va', 0)/1000:,.2f} kVA)", "Includes power factor (PF) & inductive surge"],
        ["Daily Energy Consumption", f"{sizing.get('daily_energy_kwh', 0):.2f} kWh/day", "Base load schedule requirement"],
        ["Design Target Energy (with losses)", f"{sizing.get('design_energy_kwh', 0):.2f} kWh/day", "Accounts for system losses and performance ratio"],
        ["Proposed DC Array Capacity", f"{sizing.get('total_pv_kwp', 0):.2f} kWp", f"Total PV array rating ({sizing.get('panel_qty', 0)} pcs x {sizing.get('panel_wp', 625)}Wp)"],
        ["Proposed Inverter Size", f"{sizing.get('inverter', {}).get('kw', 0):.1f} kW / {sizing.get('inverter', {}).get('kva', 0):.1f} kVA", f"Quantity: {sizing.get('inverter', {}).get('qty', 1)} unit(s)"],
    ]
    if sys_type in ("off-grid", "hybrid"):
        summary_rows.append(["Proposed Battery Storage (BESS)", f"{sizing.get('battery', {}).get('total_kwh', 0):.2f} kWh", f"{sizing.get('battery', {}).get('qty', 0)} pcs x {sizing.get('battery', {}).get('module_kwh', 14.33)} kWh modules"])
        summary_rows.append(["Days of Autonomy & DoD", f"{sizing.get('days_of_autonomy', 2.0)} Days at {sizing.get('dod', 0.8)*100:.0f}% DoD", "Includes mandatory 1.25x degradation/conversion safety factor"])

    write_table(ws1, 4, ["Design Parameter", "Calculated Specification", "Engineering Basis & Notes"], summary_rows, {1: 32, 2: 35, 3: 50})

    # ── Sheet 2: PV Array & Stringing ──
    ws2 = wb.create_sheet(title="2. PV Array & Stringing")
    format_title_banner(ws2, "Solar PV Array & Stringing Configuration", cols=4)
    
    string_info = sizing.get("stringing", {})
    pv_rows = [
        ["PV Module Power Rating", f"{sizing.get('panel_wp', 625)} Wp", "High-efficiency monocrystalline PV module"],
        ["Total Required DC Capacity", f"{sizing.get('total_pv_kwp', 0):.2f} kWp", "Sized against daily design target energy"],
        ["Total PV Modules Required", f"{sizing.get('panel_qty', 0)} pcs", f"Rule: ceil({sizing.get('total_pv_kwp', 0):.2f} kWp / {sizing.get('panel_wp', 625)/1000:.3f} kWp)"],
        ["Max Inverter Input Voltage (Vin_max)", f"{string_info.get('max_inverter_vin', 1000.0):.0f} V DC", "Maximum allowable DC voltage per string"],
        ["Module Open Circuit Voltage (Voc)", "49.28 V DC", "Module specification at Standard Test Conditions"],
        ["Max Panels per String", f"{string_info.get('max_panels_per_string', 19)} pcs", "Formula: floor(Vin_max / (Voc * (1 + K*(Tmin - 25°C))))"],
        ["Panels per MPPT", f"{string_info.get('panels_per_mppt', 26)} panels", "Allocated based on max input power per MPPT"],
        ["Total Number of Strings", f"{string_info.get('total_strings', 4)} strings", "Optimal parallel string distribution"],
        ["Operating String Voltage", f"{string_info.get('string_voltage_v', 650):.1f} V DC", "Well within inverter MPPT tracking voltage window"],
    ]
    next_row = write_table(ws2, 4, ["PV Array Parameter", "Specification", "Engineering Formula / Verification"], pv_rows, {1: 32, 2: 30, 3: 55})
    
    # ── Draw Visual MPPT Stringing Grid ──
    grid = string_info.get("stringing_grid", [])
    if grid and len(grid) > 0 and len(grid[0]) > 0:
        # Title of the Stringing Grid
        ws2.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=len(grid[0]) + 1)
        title_cell = ws2.cell(row=next_row, column=1, value="  MPPT(STRINGING) CONFIGURATION DETAIL (PANELS PER INPUT)")
        title_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(vertical="center")
        ws2.row_dimensions[next_row].height = 24
        next_row += 1
        
        # Grid Headers (MPPT 1, MPPT 2...)
        c = ws2.cell(row=next_row, column=1, value="Input \\ MPPT")
        c.font = Font(name="Calibri", size=10, bold=True, italic=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx in range(len(grid[0])):
            c = ws2.cell(row=next_row, column=col_idx + 2, value=col_idx + 1)
            c.font = bold_font
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ensure appropriate column widths for the MPPT grid columns
            col_letter = get_column_letter(col_idx + 2)
            ws2.column_dimensions[col_letter].width = 15
            
        ws2.row_dimensions[next_row].height = 20
        next_row += 1
        
        # Grid Values
        for row_idx in range(len(grid)):
            c = ws2.cell(row=next_row, column=1, value=row_idx + 1)
            c.font = bold_font
            c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx in range(len(grid[0])):
                val = grid[row_idx][col_idx]
                val_str = str(val) if val is not None else ""
                c = ws2.cell(row=next_row, column=col_idx + 2, value=val_str)
                c.font = bold_font if val is not None else data_font
                c.border = thin_border
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[next_row].height = 20
            next_row += 1

    # ── Sheet 3: Battery Storage (BESS) ──
    ws3 = wb.create_sheet(title="3. Battery Storage (BESS)")
    format_title_banner(ws3, "Battery Energy Storage System & Protection", cols=4)
    
    batt_info = sizing.get("battery", {})
    bess_rows = [
        ["Needed BESS (Base + Buffer)", f"{batt_info.get('total_kwh', 0):.3f} kWh", "Formula: [(Daily Energy * Autonomy) / DoD] * 1.25 safety factor"],
        ["Selected Battery Module", batt_info.get("type", "Dyness Stack280 14.33kWh HV Battery"), f"{batt_info.get('module_kwh', 14.33)} kWh per high-voltage module"],
        ["Number of Battery Modules", f"{batt_info.get('qty', 0)} pcs", f"Rule: ceil({batt_info.get('total_kwh', 0):.2f} / {batt_info.get('module_kwh', 14.33)})"],
        ["Number of Battery Racks/Stacks", f"{batt_info.get('stacks', 1)} Stack(s)", "Vertical stack configuration (max 9 modules per rack)"],
        ["Actual Installed Usable BESS", f"{batt_info.get('qty', 0) * batt_info.get('module_kwh', 14.33):.3f} kWh", "Total connected physical storage capacity"],
        ["Max Charge / Discharge Current", f"{batt_info.get('breaker_a', 125)/1.25:.1f} A", "Continuous operational current under peak load"],
        ["Battery DC Breaker Rating", f"{batt_info.get('breaker_a', 125):.1f} A (160A Frame)", "Rule: 1.25 * Max Charge/Discharge Current at 1000Vdc"],
        ["Main Battery Riser Cable", "50 mm² CU/PVC/PVC-Nitrile", "From battery tower/stack to main DC breaker (300/500Vdc)"],
        ["Battery Inverter Feeder Cables", "16 mm² CU/PVC/PVC-Nitrile", "4 runs for two battery inverter inputs"],
    ]
    write_table(ws3, 4, ["BESS Parameter", "Specification", "Engineering Formula / Note"], bess_rows, {1: 32, 2: 35, 3: 55})

    # ── Sheet 4: Inverter & Switchgear ──
    ws4 = wb.create_sheet(title="4. Inverter & Switchgear")
    format_title_banner(ws4, "Inverter Selection & AC Switchboard Configuration", cols=4)
    
    inv_info = sizing.get("inverter", {})
    inv_brand = inv_info.get("brand", "On-Grid / Hybrid Solar Inverter")
    inv_volt_arch = inv_info.get("voltage_architecture", "")
    inv_rows = [
        ["Selected Inverter Model", f"{inv_brand}", f"{inv_info.get('qty', 1)} unit(s) — {inv_volt_arch}" if inv_volt_arch else f"{inv_info.get('qty', 1)} unit(s)"],
        ["Inverter Power Capacity", f"{inv_info.get('kw', 50):.1f} kW ({inv_info.get('kva', 50):.1f} kVA)", "Sized against peak demand with 1.25x safety margin"],
        ["Number of Inverters", f"{inv_info.get('qty', 1)} pcs", "3-Phase / 1-Phase inverter configuration"],
        ["Total Installed AC Capacity", f"{inv_info.get('kw', 50) * inv_info.get('qty', 1):.1f} kW", "Combined continuous output"],
        ["Inverter Type & Classification", f"On-Grid / Hybrid Solar Inverter ({sys_type.upper()})", "Smart multi-MPPT bidirectional inverter with BMS CAN/RS485"],
        ["Max Output AC Current (I_ac)", f"{inv_info.get('kw', 50)*1000 / (1.732 * 400 * 0.9):.1f} A", "Formula: P_watts / (sqrt(3) * 400V * cos_phi 0.9)"],
        ["Recommended AC Breaker Rating", f"{(inv_info.get('kw', 50)*1000 / (1.732 * 400 * 0.9))*1.25:.1f} A (125A / 160A MCCB)", "Rule: 1.25 * I_ac with adjustable thermal-magnetic trip"],
        ["DC Surge Protection", "Type II 1000V DC SPDs", "Installed in PV String Junction Boxes / DCDB"],
        ["AC Surge Protection & Isolation", "Type II 440V AC SPDs & 4-Pole Isolator", "Installed in main AC Distribution Board"],
    ]
    write_table(ws4, 4, ["Inverter / AC Parameter", "Specification", "Engineering Formula / Verification"], inv_rows, {1: 32, 2: 35, 3: 55})

    # ── Sheet 5: Cable Sizing & LPS Schedule ──
    ws5 = wb.create_sheet(title="5. Cable Sizing & LPS")
    format_title_banner(ws5, "DC/AC Cable Sizing Formulas & Earthing/LPS Schedule", cols=4)
    
    cables_info = sizing.get("cables", {})
    cable_rows = [
        ["PV DC String Cable Cross-Section", f"{cables_info.get('dc_sqmm', 4)} mm² TCU/XLPO (1.5/1.5kVdc)", f"Formula: I*r*2L / VD (Total run: {cables_info.get('dc_total_m', 120):.0f} m)"],
        ["AC Main Feeder Cable", f"1 run of 4-core {cables_info.get('ac_sqmm', 25)} mm² CU/XLPE/PVC", f"V.D check: {cables_info.get('ac_vd_pct', 2.5):.2f}% ({cables_info.get('ac_vd_v', 10):.1f}V) over {cables_info.get('ac_dist_m', 100):.0f}m"],
        ["Inverter to PVDB Earthing Cable", "16 mm² CU/PVC (450/750V)", "Estimated run length: 15 m"],
        ["PVDB to Main Earth Bar Cable", "16 mm² CU/PVC (450/750V)", "Estimated run length: 40 m"],
        ["PV Rail-to-Rail Bonding Cable", "6 mm² CU/PVC (450/750V)", "Inter-module structure bonding"],
        ["Roof / Mounting Structure to PVDB", "16 mm² CU/PVC (450/750V)", "Estimated run length: 60 m"],
        ["Roof / Mounting Structure to Earthpit", "16 mm² CU/PVC (450/750V)", "Direct earth connection run: 15 m"],
        ["Battery Tower to PVDB Earth Cable", "16 mm² CU/PVC (450/750V)", "Estimated run length: 5 m"],
        ["Lightning Protection System (LPS)", "25x3mm Pure Copper Tape", "Perimeter roof ring & down conductors to earthpit: 135 m"],
    ]
    write_table(ws5, 4, ["Circuit / Earthing Connection", "Cable & Conductor Specification", "Formula / Run Length Schedule"], cable_rows, {1: 35, 2: 35, 3: 55})

    # ── Sheet 6: Complete BOQ ──
    ws6 = wb.create_sheet(title="6. Bill of Quantities (BOQ)")
    format_title_banner(ws6, "Bill of Quantities (Quantities Only — No Pricing)", cols=8)
    
    boq_headers = ["Item No.", "Category", "Description", "Quantity", "Unit", "Unit Cost", "Total Cost", "Remarks"]
    curr_row = 4
    for col_idx, h in enumerate(boq_headers, 1):
        c = ws6.cell(row=curr_row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws6.row_dimensions[curr_row].height = 26
    curr_row += 1

    curr_cat = ""
    for item in boq_items:
        cat = item.get("category", "")
        if cat != curr_cat:
            curr_cat = cat
            ws6.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
            cat_cell = ws6.cell(row=curr_row, column=1, value=f"  {curr_cat.upper()}")
            cat_cell.font = bold_font
            cat_cell.fill = section_fill
            cat_cell.alignment = Alignment(vertical="center")
            ws6.row_dimensions[curr_row].height = 22
            for c in range(1, 9):
                ws6.cell(row=curr_row, column=c).border = thin_border
            curr_row += 1

        ws6.cell(row=curr_row, column=1, value=item.get("item_no", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws6.cell(row=curr_row, column=2, value=cat).alignment = Alignment(vertical="center")
        desc_c = ws6.cell(row=curr_row, column=3, value=item.get("description", ""))
        desc_c.alignment = Alignment(vertical="center", wrap_text=True)
        ws6.cell(row=curr_row, column=4, value=item.get("quantity", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws6.cell(row=curr_row, column=5, value=item.get("unit", "")).alignment = Alignment(horizontal="center", vertical="center")
        
        # Blank price columns
        ws6.cell(row=curr_row, column=6, value="").fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws6.cell(row=curr_row, column=7, value="").fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws6.cell(row=curr_row, column=8, value=item.get("remarks", "")).alignment = Alignment(vertical="center")
        
        for c in range(1, 9):
            ws6.cell(row=curr_row, column=c).font = data_font
            ws6.cell(row=curr_row, column=c).border = thin_border
        ws6.row_dimensions[curr_row].height = 24
        curr_row += 1

    boq_widths = {1: 10, 2: 22, 3: 45, 4: 10, 5: 10, 6: 16, 7: 16, 8: 35}
    for col_idx, w in boq_widths.items():
        ws6.column_dimensions[get_column_letter(col_idx)].width = w

    out_wb = BytesIO()
    wb.save(out_wb)
    return out_wb.getvalue()

